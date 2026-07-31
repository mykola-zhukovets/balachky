#!/usr/bin/env python3
"""Впровадження україномовних текстів «на Ви» з Excel-таблиці ChatGPT.

Джерело: аркуш «Тексти Балачки» у робочій таблиці текстів (поза репозиторієм;
шлях задає змінна середовища BALACHKY_TEXTS_XLSX)
(колонки: Джерело/Ключ | Категорія | Оригінал | hy3 | Gemini | ChatGPT |
Рішення Миколи | Рекомендація | Джерело). Аркуш «English» тут НЕ чіпаємо.

Правило вибору цільового значення для КОЖНОГО рядка (канон Миколи, 22.07):
  1) непорожнє «Рішення Миколи» → беремо його;
  2) інакше непорожня «Пропозиція ChatGPT» → беремо її;
  3) інакше лишаємо Оригінал (нічого не міняємо).
Тон: звертання на «Ви»/«Вас»/«Ваш» з великої, офіційно. Колонка ChatGPT вже
такою зроблена — застосовуємо як є, вручну не переписуємо.

Захист (рядок ПРОПУСКАЄМО і пишемо в skipped):
  * ключа немає в i18n.py STRINGS["uk"] (застарілий) — пропускаємо;
  * набір плейсхолдерів {…} та %-токенів нового значення НЕ збігається з
    ПОТОЧНИМ значенням у i18n.py — пропускаємо (щоб tr(...).format() не впав, а
    парність плейсхолдерів uk/en не розсинхронилась). Порівнюємо саме з поточним
    значенням файлу (а не зі стовпцем «Оригінал», який може бути застарілим від
    старішої версії i18n) — бо парність uk↔en рахується від того, що у файлі.
  * TEST_LOCKED — ключі, чий неформальний ('ти') варіант хардкоджений у тесті як
    очікуване значення; заміна на 'Ви' зламала б тест. Рішення про оновлення
    тесту — за оркестратором/Миколою, тут НЕ чіпаємо.

Нормалізація лапок у новому значенні (house-style, test_i18n_canon вимагає лише “ ”):
  * «ялинки» та „лапки-низом“ → “ ”;
  * прямі "…" НАВКОЛО ВИДИМОГО тексту → “…”, але HTML (href="…") не чіпаємо;
  * апострофи (' ’) не чіпаємо.

Застосування:
  * i18n.py: заміна значення ключа у STRINGS["uk"] через AST. Позиції беремо в
    БАЙТАХ (ast col_offset — це utf-8-байтовий зсув у рядку), і кожну заміну
    звіряємо з ast.get_source_segment. Заміняємо лише реальні зміни (final != поточне)
    → чистий diff. Багаторядкову неявну конкатенацію згортаємо в один літерал.
  * installer/balachky.iss: рядки ukrainian.<KEY>=… у секції [CustomMessages]
    (UTF-8 BOM, CRLF — зберігаємо).

Детерміноване робиться СКРИПТОМ, а не переписуванням значень руками.

Використання:
  python scripts/apply_vy_texts.py            # застосувати
  python scripts/apply_vy_texts.py --dry-run  # лише порахувати й вивести звіт
  python scripts/apply_vy_texts.py --xlsx <шлях> --report-json <шлях>
"""
from __future__ import annotations

import argparse
import ast
import json
import re
import sys
from collections import Counter
import os
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
I18N_PATH = REPO / "fronts" / "desktop" / "i18n.py"
ISS_PATH = REPO / "installer" / "balachky.iss"
# Шлях до таблиці текстів задає той, хто запускає скрипт: особисте
# сховище автора не має місця в публічному коді.
XLSX_DEFAULT = os.environ.get("BALACHKY_TEXTS_XLSX", "")
SHEET = "Тексти Балачки"

# Ключі, які НЕ чіпаємо через хардкод неформального варіанта в тесті.
# Фікс 6: set_obsidian_hint РОЗБЛОКОВАНО — і i18n, і tests/test_ux_texts_wave.py
# уже на «Ви» («…якщо Ви нею не користуєтеся…»), тож замок застарів. Наразі порожньо;
# додавати ключ сюди лише коли тест реально хардкодить 'ти'-варіант.
TEST_LOCKED: set[str] = set()

# Плейсхолдери: іменовані {…} (в т.ч. кирилиця {дата}) + %-токени. Ширше за
# вказаний у ТЗ \{[a-z_]+\} — тобто НЕ менш захисне, ловить і кириличні поля.
_BRACE = re.compile(r"\{[^}]*\}")
_PCT = re.compile(r"%[%sdriouxXeEfgGc\d.\-+ #]*")


def placeholders(s: str) -> Counter:
    c = Counter(_BRACE.findall(s))
    c.update(_PCT.findall(s))
    return c


def normalize_quotes(s: str) -> str:
    """«»/„ → “ ”; прямі \"…\" навколо видимого тексту → “…” (HTML не чіпаємо)."""
    s = s.replace("«", "“").replace("»", "”").replace("„", "“")
    if "<" not in s and '"' in s:
        s = re.sub(r'"([^"]*)"', "“\\1”", s)
    return s


def py_str_literal(s: str) -> str:
    """Python-літерал у подвійних лапках з коректним екрануванням."""
    s = (s.replace("\\", "\\\\").replace('"', '\\"')
         .replace("\n", "\\n").replace("\r", "\\r").replace("\t", "\\t"))
    return '"' + s + '"'


def read_rows(xlsx: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[SHEET]
    rows = []
    for r in range(2, ws.max_row + 1):
        key = str(ws.cell(r, 1).value or "").strip()
        if ":" not in key:
            continue                      # рядок-розділювач секції — пропускаємо
        src, k = (p.strip() for p in key.split(":", 1))
        if src not in ("i18n.py", "installer.iss"):
            continue

        def cell(col):
            v = ws.cell(r, col).value
            return "" if v is None else str(v)

        rows.append(dict(row=r, src=src, key=k,
                         orig=cell(3), chatgpt=cell(6), mykola=cell(7)))
    return rows


def choose(row: dict):
    if row["mykola"].strip():
        return row["mykola"], "mykola"
    if row["chatgpt"].strip():
        return row["chatgpt"], "chatgpt"
    return None, "kept"


def load_uk_value_nodes(source: str) -> dict:
    tree = ast.parse(source)
    uk = None
    for node in ast.walk(tree):
        if isinstance(node, ast.Assign) and any(
                isinstance(t, ast.Name) and t.id == "STRINGS" for t in node.targets):
            for kk, vv in zip(node.value.keys, node.value.values):
                if isinstance(kk, ast.Constant) and kk.value == "uk":
                    uk = vv
    if uk is None:
        raise SystemExit("не знайдено STRINGS['uk'] в i18n.py")
    m = {}
    for kk, vv in zip(uk.keys, uk.values):
        if isinstance(kk, ast.Constant) and isinstance(kk.value, str):
            if kk.value in m:
                raise SystemExit(f"дубль ключа в uk: {kk.value}")
            m[kk.value] = vv
    return m


def apply_i18n(source: str, edits: list) -> str:
    """edits: [(value_node, new_literal_str)]. Заміна в байтах, звірка з
    get_source_segment; застосування з кінця файлу, щоб зсуви не «пливли»."""
    raw_lines = source.splitlines(keepends=True)
    line_start = [0] * (len(raw_lines) + 2)
    b = 0
    for i, line in enumerate(raw_lines, start=1):
        line_start[i] = b
        b += len(line.encode("utf-8"))
    data = source.encode("utf-8")
    spans = []
    for node, lit in edits:
        start = line_start[node.lineno] + node.col_offset
        end = line_start[node.end_lineno] + node.end_col_offset
        seg = ast.get_source_segment(source, node)
        got = data[start:end].decode("utf-8")
        if got != seg:
            raise SystemExit(f"зсув не збігся з сегментом: {got!r} != {seg!r}")
        spans.append((start, end, lit))
    for start, end, lit in sorted(spans, reverse=True):
        data = data[:start] + lit.encode("utf-8") + data[end:]
    return data.decode("utf-8")


def apply_iss(text: str, changes: dict) -> str:
    """Заміна значення після ukrainian.<KEY>= (зберігаємо CR у кінці рядка)."""
    for key, newval in changes.items():
        pat = re.compile(r"(?m)^(ukrainian\." + re.escape(key) + r"=).*?(\r?)$")
        text, n = pat.subn(lambda m: m.group(1) + newval + m.group(2), text, count=1)
        if n != 1:
            raise SystemExit(f"не знайдено рядок ukrainian.{key}= в balachky.iss")
    return text


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=XLSX_DEFAULT)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--report-json", default="")
    args = ap.parse_args()

    rows = read_rows(args.xlsx)
    i18n_source = I18N_PATH.read_text(encoding="utf-8")
    uk_nodes = load_uk_value_nodes(i18n_source)
    uk_cur = {k: ast.literal_eval(v) for k, v in uk_nodes.items()}

    report = dict(
        total=len(rows),
        sel_mykola=0, sel_chatgpt=0, sel_kept=0,
        i18n_changed=0, i18n_noop=0,
        installer_changed=0, installer_noop=0,
        applied_mykola=0, applied_chatgpt=0, kept_original=0,
        quotes_normalized=0, excel_orig_stale=0,
        skipped=[],
    )
    i18n_edits = []            # (node, literal)
    iss_changes = {}           # key -> newval
    changed_keys = []          # (key, source) applied

    for row in rows:
        chosen, source = choose(row)
        if source == "mykola":
            report["sel_mykola"] += 1
        elif source == "chatgpt":
            report["sel_chatgpt"] += 1
        else:
            report["sel_kept"] += 1

        key, src = row["key"], row["src"]

        if src == "i18n.py":
            cur = uk_cur.get(key)
            if cur is None:
                report["skipped"].append(
                    dict(row=row["row"], key=key, reason="немає ключа в i18n.py"))
                continue
            if row["orig"] and row["orig"] != cur:
                report["excel_orig_stale"] += 1
            if source == "kept":
                report["kept_original"] += 1
                continue
            if key in TEST_LOCKED:
                report["skipped"].append(dict(
                    row=row["row"], key=key,
                    reason="TEST_LOCKED: тест хардкодить 'ти'-варіант "
                           "(tests/test_ux_texts_wave.py)"))
                continue
            final = normalize_quotes(chosen)
            if final != chosen:
                report["quotes_normalized"] += 1
            if placeholders(final) != placeholders(cur):
                report["skipped"].append(dict(
                    row=row["row"], key=key,
                    reason="розсинхрон плейсхолдерів проти поточного i18n"))
                continue
            if final == cur:
                report["i18n_noop"] += 1
                report["kept_original"] += 1
                continue
            i18n_edits.append((uk_nodes[key], py_str_literal(final)))
            report["i18n_changed"] += 1
            report["applied_mykola" if source == "mykola" else "applied_chatgpt"] += 1
            changed_keys.append((key, source))

        else:  # installer.iss
            if source == "kept":
                report["kept_original"] += 1
                continue
            final = normalize_quotes(chosen)
            if final != chosen:
                report["quotes_normalized"] += 1
            # плейсхолдери інсталятора не звіряємо з i18n; %-токенів у цих рядках нема
            if placeholders(final) != placeholders(row["orig"]):
                report["skipped"].append(dict(
                    row=row["row"], key=key,
                    reason="розсинхрон плейсхолдерів installer проти Оригіналу"))
                continue
            iss_changes[key] = final
            changed_keys.append((key, source))

    # installer: визначити реальні зміни проти поточного файлу.
    # applied_* для installer-ключів рахуємо ТУТ (у головному циклі не рахували).
    iss_source = ISS_PATH.read_text(encoding="utf-8-sig")
    src_by_key = dict(changed_keys)
    real_iss = {}
    for key, newval in iss_changes.items():
        m = re.search(r"(?m)^ukrainian\." + re.escape(key) + r"=(.*?)\r?$", iss_source)
        curval = m.group(1) if m else None
        if curval == newval:
            report["installer_noop"] += 1
            report["kept_original"] += 1
        else:
            real_iss[key] = newval
            report["installer_changed"] += 1
            s = src_by_key[key]
            report["applied_mykola" if s == "mykola" else "applied_chatgpt"] += 1

    if not args.dry_run:
        if i18n_edits:
            new_i18n = apply_i18n(i18n_source, i18n_edits)
            I18N_PATH.write_text(new_i18n, encoding="utf-8", newline="")
        if real_iss:
            new_iss = apply_iss(iss_source, real_iss)
            ISS_PATH.write_text(new_iss, encoding="utf-8-sig", newline="")

    report["skipped_count"] = len(report["skipped"])
    out = json.dumps(report, ensure_ascii=False, indent=1)
    if args.report_json:
        Path(args.report_json).write_text(out, encoding="utf-8")
    # у stdout лише лічильники (без кириличних значень) — безпечно для cp1251-консолі
    safe = {k: v for k, v in report.items() if k != "skipped"}
    safe["skipped_keys"] = [s["key"] for s in report["skipped"]]
    print(json.dumps(safe, ensure_ascii=True, indent=1))
    return 0


if __name__ == "__main__":
    sys.exit(main())
